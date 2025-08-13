from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from .forms import QueryForm, QueryChangePasswordForm, QueryParameterForm, QueryEditForm
from .models import Query, QueryParameter, RequestLog
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib.auth.hashers import check_password, make_password
import psycopg2
from django.utils import timezone
from django.contrib.auth.decorators import user_passes_test
import time
import json
import openpyxl
from openpyxl.utils import get_column_letter
from .dynamic_forms import build_dynamic_run_form
from django.template import Template, Context



@user_passes_test(lambda u: u.is_superuser)
def create_query(request):
    if request.method == 'POST':
        form = QueryForm(request.POST)
        if form.is_valid():
            query = form.save(commit=False)
            query.user = request.user
            query.password = make_password(query.password)
            query.save()
            return redirect('create_query_parameter', query_id=query.pk)
    else:
        form = QueryForm()
    
    return render(request, 'create_query.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def create_query_parameter(request, query_id):
    query = get_object_or_404(Query, id=query_id)

    if request.method == 'POST':
        form = QueryParameterForm(request.POST)
        if form.is_valid():
            param = form.save(commit=False)
            param.query = query
            param.save()
            return redirect('create_query_parameter', query_id=query.id)  
    else:
        form = QueryParameterForm()
    return render(request, 'create_query_parameter.html', {'form': form, 'query': query})


@login_required
def queries_all(request):
    queries = Query.objects.all()
    return render(request, 'queries_all.html', {'queries': queries})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def queries_my(request):
    queries = Query.objects.filter(user=request.user)
    return render(request, 'queries_my.html', {'queries': queries})

@login_required
@user_passes_test(lambda u: u.is_superuser)
@require_POST
def delete_query(request, query_uuid):
    Query.objects.filter(uuid=query_uuid, user=request.user).delete()
    return redirect('/queries/queries_my')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def edit_query(request, query_uuid):
    query = get_object_or_404(Query, uuid=query_uuid)
    params = query.parameters.all()

    if request.method == "POST":
        form = QueryEditForm(request.POST, instance=query)
        if form.is_valid():
            form.save()
            return redirect("/queries/queries_my")
    else:
        form = QueryEditForm(instance=query)

    return render(request, "edit_query.html", {
        "query": query,
        "form": form,
        "params": params
    })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def change_password(request, query_uuid):
    query = get_object_or_404(Query, uuid=query_uuid, user=request.user)

    if request.method == 'POST':
        form = QueryChangePasswordForm(request.POST)
        if form.is_valid():
            old = form.cleaned_data['old_password']
            new = form.cleaned_data['new_password']
            print('start', new, old)
            if not query.password:
                query.password = make_password(new)
                query.save()
                return redirect('edit', query_uuid=query.uuid)
            else:
                if not check_password(old, query.password):
                    form.add_error('old_password', 'Wrong Old Password')
                else:
                    query.password = make_password(new)
                    query.save()
                    return redirect('edit', query_uuid=query.uuid)
    else:
        form = QueryChangePasswordForm()
    return render(request, 'change_password.html', {'form': form, 'query': query})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def reset_password(request, query_uuid):
    query = get_object_or_404(Query, uuid=query_uuid, user=request.user)

    if request.method == 'POST':
        account_password = request.POST.get('account_password', '')
        if request.user.check_password(account_password):
            query.password = make_password('')
            query.save()
            return redirect('edit', query_uuid=query.uuid)

    return redirect('edit', query_uuid=query.uuid)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def manage_query_params(request, query_uuid):
    query = get_object_or_404(Query, uuid=query_uuid)
    params = query.parameters.all()

    param_id = request.GET.get("edit")
    if param_id:
        param = get_object_or_404(QueryParameter, id=param_id, query=query)
    else:
        param = None

    if request.method == "POST":
        form = QueryParameterForm(request.POST, instance=param)
        if form.is_valid():
            param_obj = form.save(commit=False)
            param_obj.query = query
            param_obj.save()
            return redirect("manage_query_params", query_uuid=query.uuid)
    else:
        form = QueryParameterForm(instance=param)

    return render(request, "manage_query_params.html", {
        "query": query,
        "params": params,
        "form": form,
        "editing": bool(param_id),
    })

def delete_query_param(request, param_id):
    param = get_object_or_404(QueryParameter, id=param_id)
    query_uuid = param.query.uuid
    param.delete()
    return redirect("manage_query_params", query_uuid=query_uuid)

@login_required
def run_query(request, query_uuid):
    query = get_object_or_404(Query, uuid=query_uuid)
    parameters = query.parameters.all()
    DynamicRunForm = build_dynamic_run_form(parameters)

    session_key = f'query_authenticated_{str(query.uuid)}'
    if query.password and not request.session.get(session_key, False):
        if request.method == 'POST' and 'query_password' in request.POST:
            if check_password(request.POST.get('query_password'), query.password):
                request.session[session_key] = True
                return redirect(request.path)
            else:
                error = "Wrong Password"
                return render(request, 'query_password.html', {'query': query, 'error': error})
        else:
            return render(request, 'query_password.html', {'query': query})

    if request.method == 'POST':
        form = DynamicRunForm(request.POST)
        if form.is_valid():
            context_data = {}
            for param in parameters:
                if param.type == 'date':
                    date_value = form.cleaned_data.get(param.name)
                    if date_value:
                        context_data[param.name] = date_value.isoformat()
                    else:
                        None
                elif param.type == 'number':
                    context_data[f"{param.name}_min"] = form.cleaned_data.get(f"{param.name}_min")
                    context_data[f"{param.name}_max"] = form.cleaned_data.get(f"{param.name}_max")
                else:
                    context_data[param.name] = form.cleaned_data.get(param.name)

            t = Template(query.template)
            c = Context(context_data)
            sql = t.render(c)
        
            try:
                start_time = time.time()
                conn = psycopg2.connect(query.db_dsn)
                cur = conn.cursor()
                cur.execute(sql)  
                result = cur.fetchall()
                columns = [desc[0] for desc in cur.description]
                cur.close()
                conn.close()
                end_time = time.time()

                data_for_log = [dict(zip(columns, row)) for row in result]

                log = RequestLog.objects.create(
                    query=query,
                    request=sql,
                    response=json.dumps(data_for_log, default=str),
                    response_code=0,
                    response_time=end_time - start_time
                )

                query.execution_count = (query.execution_count or 0) + 1
                query.last_executed_at = timezone.now()
                query.save()

            except Exception as e:
                log = RequestLog.objects.create(
                    query=query,
                    request=sql,
                    response=str(e),
                    response_code=1,
                    response_time=0
                )
                return render(request, 'query_result.html', {
                    'query': query,
                    'form': form,
                    'error': str(e),
                    'request_log_id': log.id
                })

            return render(request, 'query_result.html', {
                'query': query,
                'form': form,
                'result': result,
                'columns': columns,
                'sql': sql,
                'error': None,
                'request_log_id': log.id
            })

    else:
        form = DynamicRunForm()

    return render(request, 'queryparameter_form.html', {'form': form, 'query': query})



def download_excel(request, request_log_id):
    log = get_object_or_404(RequestLog, id=request_log_id)
    try:
        data = json.loads(log.response) 
    except (json.JSONDecodeError, TypeError):
        return HttpResponse("Invalid or empty result", status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Query Result"

    if data:
        headers = data[0].keys()
        for idx, header in enumerate(headers, start=1):
            ws[f"{get_column_letter(idx)}1"] = header

        for row_idx, row in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws[f"{get_column_letter(col_idx)}{row_idx}"] = row.get(header)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="query_result_{log.id}.xlsx"'
    wb.save(response)
    return response